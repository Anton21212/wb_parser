import logging
from typing import List

import openpyxl
from pydantic import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from core.serializers import CardSerializer
from core.validators import Card
from utils.wildberries import get_cards_info, CardInfo

logger = logging.getLogger('django')


class WbApiView(APIView):
    @staticmethod
    def post(request):
        serializer = CardSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        articles_file = serializer.validated_data.get('articles_file', None)
        article = serializer.validated_data.get('article', None)

        if articles_file is not None:
            try:
                card_ids = _get_card_ids_from_xlsx(articles_file)
            except Exception as err:
                logger.info(f'WbApiView.post; get_card_ids_from_xlsx; err: {err}')
                return Response("Invalid file", status=400)
        elif article is not None:
            card_ids = [article]
        else:
            return Response("Bad request", status=400)
        if not card_ids:
            return Response(data=[], status=200)

        try:
            cards_info = get_cards_info(card_ids=card_ids)
        except Exception as err:
            logger.warning(f'WbApiView.post; get_cards_info; err: {err}')
            return Response("Wildberries not available", status=500)
        response_data = _parse_cards_info(cards_info=cards_info)
        if len(response_data) == 1:
            response_data = response_data[0]
        return Response(response_data)


def _parse_cards_info(cards_info: List[CardInfo]) -> List[dict]:
    result = []
    for card_info in cards_info:
        if card_info.error:
            card_data = {
                'article': card_info.id,
                'error': card_info.error
            }
            result.append(card_data)
            continue
        try:
            card_obj = Card.parse_obj(card_info.info)
            card_data = card_obj.dict()
        except ValidationError:
            card_data = {
                'article': card_info.id,
                'error': 'Unparsed'
            }
        result.append(card_data)
    return result


def _get_card_ids_from_xlsx(file) -> List[int]:
    wb = openpyxl.load_workbook(file)
    first_sheet = wb.sheetnames[0]
    worksheet = wb[first_sheet]
    card_ids = []
    for row in range(1, worksheet.max_row + 1):
        cell_name = f"A{row}"
        card_id = worksheet[cell_name].value
        if isinstance(card_id, int):
            card_ids.append(card_id)
    return card_ids
